from abc import ABC
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response

from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from django.http import HttpResponse
from dormitory.models import Faculty
from datetime import date
import uuid


def upload_path(prefix, instance, filename):
    ext = filename.split('.')[-1]
    today = date.today().strftime('%Y/%m/%d')
    filename = f'{uuid.uuid4().hex[:16]}.{ext}'
    return '{prefix}/{today}/{filename}'.format(
        prefix=prefix,
        today=today,
        filename=filename)


def upload_photo(instance, filename):
    return upload_path('student/diploma', instance, filename)


class CustomPagination(PageNumberPagination, ABC):
    page_size = 15
    max_page_size = 100
    page_size_query_param = 'page_size'

    def get_page_size(self, request):
        page_size = request.query_params.get('page_size', self.page_size)
        if page_size is not None:
            return min(int(page_size), self.max_page_size)
        return self.page_size

    def get_paginated_response(self, data):
        return Response({
            'pagination': {
                'next': self.get_next_link(),
                'previous': self.get_previous_link(),
                'total': self.page.paginator.count,
                'page_size': int(self.request.GET.get('page_size', self.page_size)),
                'current_page_number': self.page.number,
                'total_pages': self.page.paginator.num_pages,
            },
            'results': data
        })

        # 'links': {
        #     'next': self.get_next_link(),
        #     'previous': self.get_previous_link()
        # },
        # 'total': self.page.paginator.count,
        # 'page_size': int(self.request.GET.get('page_size', self.page_size)),
        # 'current_page_number': self.page.number,
        # 'total_pages': self.page.paginator.num_pages,
        # 'results': data


def export_data(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    faculty_queryset = Faculty.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-student.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Movies'

    # Define the titles for columns
    columns = [
        'id',
        'name',
        'Description',
        'Length',
        'Rating',
        'Price',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for movie in faculty_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            movie.pk,
            movie.title,
            movie.description,
            movie.length_in_minutes,
            movie.rating,
            movie.price,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response