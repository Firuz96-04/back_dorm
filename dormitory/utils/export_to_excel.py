from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.utils import get_column_letter


def export_data(request, totals, book_name):
    faculty_queryset = request

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-{book_name}.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
        book_name=book_name
    )
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = book_name.title()

    columns = [
        ('#', 5),
        ('Факультет', 20),
        ('Кол. студентов', 15),
        ('Заселены', 15),

    ]
    row_num = 1

    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    for item, faculty in enumerate(faculty_queryset):
        row_num += 1

        row = [
            item++1,
            faculty.name,
            faculty.student_count,
            faculty.booking_count
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    last_row = worksheet.max_row
    worksheet.append(['Общее количество', '', totals['student_total'], totals['book_total']])
    worksheet.merge_cells(f'A{last_row+1}:B{last_row+1}')

    workbook.save(response)

    return response
